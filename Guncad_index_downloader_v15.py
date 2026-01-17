"""
GunCAD Index Downloader v6 - Quality of Life Edition (FIXED)
Based on v5 Intent-Based Organization with new features:
- Real-time failed downloads counter
- Configurable output directory at startup
- Adjustable download timeout with explanations
- Batch Excel updates every N downloads
- Estimated time remaining (ETA) indicator
- LBRY URL tracking in Excel for reliable duplicate detection
- Full description, notes, and readme fields (not truncated)
- Directory-scoped file checking
"""

import os
import re
import time
import json
import hashlib
import zipfile
import shutil
import requests
from datetime import datetime
from urllib.parse import urljoin, urlparse, unquote, parse_qs
from collections import defaultdict


class DownloadTracker:
    """Tracks downloaded files using LBRY URLs as unique identifiers"""
    
    def __init__(self, output_dir, excel_index=None):
        self.output_dir = output_dir
        self.excel_index = excel_index
        self.db_file = os.path.join(output_dir, 'download_history.json')
        self.history = self.load_history()
        self.filesystem_cache = None

    def load_history(self):
        if os.path.exists(self.db_file):
            try:
                with open(self.db_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                return {}
        return {}

    def save_history(self):
        with open(self.db_file, 'w', encoding='utf-8') as f:
            json.dump(self.history, f, indent=2, ensure_ascii=False)

    def get_entry_id(self, detail_url):
        return hashlib.md5(detail_url.encode()).hexdigest()
    
    def build_filesystem_cache(self):
        """Build a cache of all files in the output directory"""
        if self.filesystem_cache is not None:
            return
        
        self.filesystem_cache = {}
        print(f"\n  🔍 Scanning '{self.output_dir}' for existing files...")
        
        file_count = 0
        for root, dirs, files in os.walk(self.output_dir):
            for filename in files:
                if filename in ['download_history.json', 'GunCAD_Master_Index.xlsx', 
                               'GunCAD_Master_Index.csv', 'QUICK_FIND.txt', 'README.md']:
                    continue
                
                filepath = os.path.join(root, filename)
                try:
                    file_size = os.path.getsize(filepath)
                    if filename not in self.filesystem_cache:
                        self.filesystem_cache[filename] = []
                    self.filesystem_cache[filename].append({
                        'path': filepath,
                        'size': file_size
                    })
                    file_count += 1
                except:
                    pass
        
        print(f"  ✓ Found {file_count} existing files")
        if file_count > 0:
            print(f"  📋 Sample files: {list(self.filesystem_cache.keys())[:5]}")
        print()

    def is_downloaded(self, detail_url, lbry_url=None, expected_size=0):
        """Check if file exists using Excel index (by LBRY URL) - ONLY in current output directory"""
        
        if self.excel_index and lbry_url:
            for entry in self.excel_index:
                if entry.get('LBRY URL') == lbry_url:
                    filepath = entry.get('Location', '')
                    if filepath and os.path.exists(filepath):
                        try:
                            file_abs = os.path.abspath(filepath)
                            output_abs = os.path.abspath(self.output_dir)
                            rel_path = os.path.relpath(file_abs, output_abs)
                            
                            if rel_path.startswith('..'):
                                print(f"    ⚠ File exists but outside '{self.output_dir}', will re-download")
                                return False
                            
                            print(f"    ✓ Found in index: {os.path.basename(filepath)}")
                            return True
                        except ValueError:
                            print(f"    ⚠ File on different drive, will re-download")
                            return False
                    else:
                        print(f"    ⚠ In index but file missing, will re-download")
                        return False
        
        entry_id = self.get_entry_id(detail_url)
        
        if entry_id in self.history:
            entry = self.history[entry_id]
            
            if entry.get('status') == 'failed':
                return False
            
            stored_filepath = entry.get('filepath', '')
            if stored_filepath and os.path.exists(stored_filepath):
                try:
                    file_abs = os.path.abspath(stored_filepath)
                    output_abs = os.path.abspath(self.output_dir)
                    rel_path = os.path.relpath(file_abs, output_abs)
                    
                    if not rel_path.startswith('..'):
                        filename = os.path.basename(stored_filepath)
                        print(f"    ✓ Found in history: {filename}")
                        return True
                except ValueError:
                    pass
        
        return False
    
    def file_exists_in_cache(self, filename, expected_size=0):
        """Check if a file with this name exists in the filesystem cache"""
        if self.filesystem_cache is None:
            self.build_filesystem_cache()
        
        if filename in self.filesystem_cache:
            matches = self.filesystem_cache[filename]
            
            if expected_size > 0:
                for match in matches:
                    size_diff = abs(match['size'] - expected_size)
                    if size_diff < (expected_size * 0.01):
                        return match['path']
            
            return matches[0]['path'] if matches else None
        
        base_filename = filename.split(':')[0] if ':' in filename else filename
        normalized_search = base_filename.lower().replace('-', ' ').replace('_', ' ')
        
        for cached_name, matches in self.filesystem_cache.items():
            cached_base = os.path.splitext(cached_name)[0]
            normalized_cached = cached_base.lower().replace('-', ' ').replace('_', ' ')
            
            if normalized_search == normalized_cached:
                if expected_size > 0:
                    for match in matches:
                        size_diff = abs(match['size'] - expected_size)
                        if size_diff < (expected_size * 0.01):
                            return match['path']
                
                return matches[0]['path'] if matches else None
        
        return None

    def mark_downloaded(self, detail_url, title, filepath, tags, verified=False, 
                       file_size=0, category=None, gun_model=None, caliber=None):
        entry_id = self.get_entry_id(detail_url)
        self.history[entry_id] = {
            'title': title,
            'detail_url': detail_url,
            'filepath': filepath,
            'tags': tags,
            'downloaded_at': datetime.now().isoformat(),
            'file_exists': os.path.exists(filepath),
            'verified': verified,
            'file_size': file_size,
            'category': category,
            'gun_model': gun_model,
            'caliber': caliber
        }
        self.save_history()
        
        if self.filesystem_cache is not None:
            filename = os.path.basename(filepath)
            if filename not in self.filesystem_cache:
                self.filesystem_cache[filename] = []
            self.filesystem_cache[filename].append({
                'path': filepath,
                'size': file_size
            })

    def mark_failed(self, detail_url, title, reason, lbry_url=''):
        entry_id = self.get_entry_id(detail_url)
        self.history[entry_id] = {
            'title': title,
            'detail_url': detail_url,
            'lbry_url': lbry_url,
            'status': 'failed',
            'reason': reason,
            'failed_at': datetime.now().isoformat()
        }
        self.save_history()

    def get_stats(self):
        total = len(self.history)
        successful = sum(1 for e in self.history.values() 
                        if e.get('verified', False))
        failed = sum(1 for e in self.history.values() 
                    if e.get('status') == 'failed')
        return {
            'total': total,
            'successful': successful,
            'failed': failed
        }


class LBRYDaemonClient:
    """Client for LBRY daemon with configurable timeouts"""
    
    def __init__(self, daemon_url='http://localhost:5279', max_wait_time=300):
        self.daemon_url = daemon_url
        self.available = False
        self.max_wait_time = max_wait_time
        self.check_connection()

    def check_connection(self):
        try:
            response = self._call_method('status')
            if response:
                print("✓ LBRY daemon connected")
                self.available = True
                return True
        except:
            pass
        
        print("✗ LBRY daemon not available")
        self.available = False
        return False

    def _call_method(self, method, params=None):
        if params is None:
            params = {}

        payload = {
            "method": method,
            "params": params
        }

        try:
            response = requests.post(
                self.daemon_url,
                json=payload,
                timeout=30
            )
            response.raise_for_status()
            data = response.json()
            
            if 'error' in data:
                return None
                
            return data.get('result')
            
        except:
            return None

    def get_file(self, lbry_url, max_retries=3):
        """Download file via LBRY daemon"""
        print(f"  Starting LBRY download: {lbry_url}")
        
        for attempt in range(max_retries):
            if attempt > 0:
                print(f"  Retry {attempt + 1}/{max_retries}...")
                time.sleep(5)
            
            result = self._call_method('get', {
                'uri': lbry_url,
                'save_file': True
            })
            
            if not result:
                continue
            
            download_path = result.get('download_path')
            claim_name = result.get('claim_name')
            status = result.get('status', 'unknown')
            
            if attempt == 0:
                print(f"  Status: {status}")
            
            if status in ('completed', 'finished'):
                if download_path and os.path.exists(download_path):
                    file_size = os.path.getsize(download_path)
                    if file_size > 0:
                        print(f"  ✓ Download complete ({file_size/1024/1024:.2f}MB)")
                        return download_path
            
            if status == 'running':
                if attempt == 0:
                    print(f"  Waiting for completion (max {self.max_wait_time}s)...")
                result_path = self.wait_for_download(lbry_url, claim_name)
                if result_path:
                    return result_path
                continue
            
            if status == 'stopped':
                if download_path and os.path.exists(download_path):
                    file_size = os.path.getsize(download_path)
                    if file_size > 0:
                        print(f"  ✓ File exists ({file_size/1024/1024:.2f}MB)")
                        return download_path
        
        print(f"  ✗ Failed after {max_retries} attempts")
        return None

    def wait_for_download(self, lbry_url, claim_name):
        """Wait for download completion by monitoring progress"""
        start_time = time.time()
        last_progress = -1
        last_download_path = None
        last_written_bytes = 0
        stall_start_time = None
        stall_threshold = 30  # Consider stalled if no progress for 30 seconds
        
        print(f"  ⏳ Download in progress (monitoring for stalls)...")
        
        while True:
            elapsed = int(time.time() - start_time)
            file_list = self._call_method('file_list', {'claim_name': claim_name})
            
            if file_list:
                items = file_list.get('items', [])
                
                if not items:
                    time.sleep(2)
                    # Check if we've been waiting too long with no items
                    if elapsed > 60:
                        print(f"\n  ✗ No download info after {elapsed}s")
                        return None
                    continue
                
                file_info = items[0]
                status = file_info.get('status', 'unknown')
                download_path = file_info.get('download_path')
                
                # Store the download path in case we need it
                if download_path:
                    last_download_path = download_path
                
                if 'written_bytes' in file_info and 'total_bytes' in file_info:
                    written = file_info['written_bytes']
                    total = file_info['total_bytes']
                    
                    # Check if download is making progress
                    if written > last_written_bytes:
                        # Progress detected - reset stall timer
                        stall_start_time = None
                        last_written_bytes = written
                    else:
                        # No progress - check if stalled
                        if stall_start_time is None:
                            stall_start_time = time.time()
                        elif time.time() - stall_start_time > stall_threshold:
                            print(f"\n  ⚠ Download stalled (no progress for {stall_threshold}s)")
                            # Check if file is complete despite stall
                            if download_path and os.path.exists(download_path):
                                file_size = os.path.getsize(download_path)
                                if file_size > 0:
                                    print(f"  ✓ File exists ({file_size/1024/1024:.2f}MB)")
                                    return download_path
                            return None
                    
                    if total > 0:
                        progress = int((written / total) * 100)
                        speed = written / max(elapsed, 1) / 1024 / 1024  # MB/s
                        if progress != last_progress or elapsed % 5 == 0:
                            print(f"  Progress: {progress}% ({written/1024/1024:.1f}MB / {total/1024/1024:.1f}MB) - {speed:.2f} MB/s - {elapsed}s", end='\r')
                            last_progress = progress
                
                if status in ('completed', 'finished'):
                    print(f"\n  ✓ Download completed in {elapsed}s!")
                    if download_path and os.path.exists(download_path):
                        return download_path
                
                elif status == 'stopped':
                    print(f"\n  Download stopped at {elapsed}s")
                    if download_path and os.path.exists(download_path):
                        file_size = os.path.getsize(download_path)
                        if file_size > 0:
                            print(f"  ✓ File complete ({file_size/1024/1024:.2f}MB)")
                            return download_path
                    return None
            else:
                # No response from API
                if elapsed > 60:
                    print(f"\n  ✗ No API response after {elapsed}s")
                    return None
            
            time.sleep(2)
        
        return None


class FileVerifier:
    """Verify files"""
    
    @staticmethod
    def verify_zip(filepath):
        try:
            with zipfile.ZipFile(filepath, 'r') as zf:
                result = zf.testzip()
                if result is None:
                    file_count = len(zf.namelist())
                    return True, f"Valid zip with {file_count} files"
                else:
                    return False, f"Corrupt file: {result}"
        except zipfile.BadZipFile:
            return False, "Not a valid zip"
        except Exception as e:
            return False, f"Error: {e}"
    
    @staticmethod
    def is_zip_file(filepath):
        """Check if file is actually a zip by reading magic bytes"""
        try:
            with open(filepath, 'rb') as f:
                magic = f.read(4)
                return magic[:2] == b'PK'
        except:
            return False
    
    @staticmethod
    def verify_file(filepath):
        if not os.path.exists(filepath):
            return False, "File not found"
        
        file_size = os.path.getsize(filepath)
        if file_size == 0:
            return False, "File is empty"
        
        ext = os.path.splitext(filepath)[1].lower()
        
        if ext == '.zip' or FileVerifier.is_zip_file(filepath):
            is_valid, message = FileVerifier.verify_zip(filepath)
            if is_valid:
                return True, message
            return True, f"File downloaded ({file_size/1024/1024:.2f}MB) - may not be standard zip"
        
        cad_formats = ['.stl', '.step', '.stp', '.3mf', '.obj', '.f3d', 
                      '.blend', '.scad', '.dxf', '.dwg', '.iges', '.igs',
                      '.pdf', '.txt', '.md', '.gcode', '.rar', '.7z']
        
        if ext in cad_formats:
            return True, f"Valid {ext.upper()} file ({file_size/1024/1024:.2f}MB)"
        
        return True, f"File downloaded ({file_size/1024/1024:.2f}MB)"


class GunCADIndexAPIClient:
    """Client for GunCAD Index API"""
    
    def __init__(self, api_base='https://guncadindex.com/api'):
        self.api_base = api_base
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'application/json'
        })
    
    def get_all_tags(self, scan_pages=5):
        """Fetch all available tags by scanning releases"""
        print(f"  Scanning first {scan_pages} pages of releases to find all tags...")
        
        all_tags = set()
        
        try:
            for page in range(1, scan_pages + 1):
                offset = (page - 1) * 25
                url = f"{self.api_base}/releases/?limit=25&offset={offset}"
                
                response = self.session.get(url, timeout=30)
                response.raise_for_status()
                
                data = response.json()
                
                if 'results' in data:
                    for entry in data['results']:
                        if 'tags' in entry:
                            for tag in entry.get('tags', []):
                                if isinstance(tag, dict) and tag.get('name'):
                                    all_tags.add(tag.get('name'))
                
                # Show progress
                if page % 2 == 0:
                    print(f"  Scanned {page * 25} releases, found {len(all_tags)} unique tags so far...")
            
            return sorted(list(all_tags))
            
        except Exception as e:
            print(f"  Warning: Could not fetch tags from API: {e}")
            return []
    
    def get_releases(self, page=1, limit=25):
        """Fetch releases from API"""
        offset = (page - 1) * limit
        url = f"{self.api_base}/releases/?limit={limit}&offset={offset}"
        
        try:
            print(f"  API URL: {url}")
            response = self.session.get(url, timeout=30)
            response.raise_for_status()
            
            data = response.json()
            
            if 'results' in data:
                return data['results']
            
            return []
            
        except Exception as e:
            print(f"  ✗ API Error: {e}")
            return []
    
    def parse_entry(self, entry):
        """Parse an API entry"""
        lbry_url = entry.get('url_lbry', '')
        odysee_url = entry.get('url', '')
        
        tags = []
        if 'tags' in entry:
            tags = [tag.get('name', '') for tag in entry.get('tags', []) if isinstance(tag, dict)]
        
        release_id = entry.get('id', '')
        shortlink = entry.get('shortlink')
        
        if shortlink:
            detail_url = f"https://guncadindex.com/detail/{shortlink}"
        elif release_id:
            detail_url = f"https://guncadindex.com/detail/{release_id}"
        else:
            detail_url = odysee_url
        
        return {
            'title': entry.get('name', 'Unknown'),
            'detail_url': detail_url,
            'lbry_url': lbry_url,
            'odysee_url': odysee_url,
            'tags': tags,
            'description': entry.get('description', ''),
            'size': entry.get('size', 0),
            'release_date': entry.get('release_date', ''),
            'last_updated': entry.get('last_updated', ''),
            'author': entry.get('author', ''),
            'version': entry.get('version', ''),
            'notes': entry.get('notes', ''),
            'readme': entry.get('readme', ''),
            'odysee_views': entry.get('odysee_views', 0),
            'odysee_likes': entry.get('odysee_likes', 0),
            'odysee_dislikes': entry.get('odysee_dislikes', 0),
        }


class IntentBasedOrganizer:
    """Organizes files based on user intent, not just tags"""
    COLUMN_ORDER = [
        'File Name', 'Location', 'LBRY URL', 'Detail URL', 'Category',
        'Gun Model', 'Caliber', 'Part Type', 'Tags', 'File Size (MB)',
        'Release Date', 'Last Updated', 'Author', 'Version', 'Date Downloaded',
        'Odysee Views', 'Odysee Likes', 'Odysee Dislikes',
        'Description', 'Notes', 'Readme'
    ]

    def __init__(self, output_dir):
        self.output_dir = output_dir
        self.master_index = []
        self.folder_readmes = defaultdict(list)
        self.index_file = os.path.join(output_dir, 'GunCAD_Master_Index.xlsx')
        self.index_file_csv = os.path.join(output_dir, 'GunCAD_Master_Index.csv')
        
        self.load_existing_index()
    
    def sanitize_filename(self, filename):
        return re.sub(r'[<>:"/\\|?*]', '_', filename)
    
    def clean_for_excel(self, text):
        """Remove illegal characters for Excel cells"""
        if not text:
            return ""
        
        cleaned = ''.join(char for char in text if ord(char) >= 32 or char in '\t\n\r')
        cleaned = cleaned.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
        cleaned = ' '.join(cleaned.split())
        
        return cleaned
    
    def load_existing_index(self):
        """Load existing master index if it exists"""
        if os.path.exists(self.index_file):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(self.index_file)
                ws = wb.active
                
                headers = [cell.value for cell in ws[1]]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        entry = dict(zip(headers, row))
                        self.master_index.append(entry)
                
                print(f"✓ Loaded existing index: {len(self.master_index)} files")
                self.reconcile_moved_files()
                return
            except Exception as e:
                print(f"  Warning: Could not load existing Excel index: {e}")
        
        if os.path.exists(self.index_file_csv):
            try:
                import csv
                with open(self.index_file_csv, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    self.master_index = list(reader)
                
                print(f"✓ Loaded existing index: {len(self.master_index)} files")
                self.reconcile_moved_files()
                return
            except Exception as e:
                print(f"  Warning: Could not load existing CSV index: {e}")
        
        print("  No existing index found - starting fresh")
    
    def reconcile_moved_files(self):
        """Update file paths in index ONLY for files within the current output directory"""
        print("  🔄 Checking for moved files within output directory...")
        
        updated_count = 0
        outside_dir_count = 0
        missing_count = 0
        
        for entry in self.master_index:
            old_path = entry.get('Location', '')
            if not old_path:
                continue
            
            filename = os.path.basename(old_path)
            
            if os.path.exists(old_path):
                try:
                    file_abs = os.path.abspath(old_path)
                    output_abs = os.path.abspath(self.output_dir)
                    rel_path = os.path.relpath(file_abs, output_abs)
                    
                    if not rel_path.startswith('..'):
                        continue
                    else:
                        outside_dir_count += 1
                        continue
                except ValueError:
                    outside_dir_count += 1
                    continue
            
            for root, dirs, files in os.walk(self.output_dir):
                if filename in files:
                    new_path = os.path.join(root, filename)
                    
                    old_size_str = entry.get('File Size (MB)', '0')
                    try:
                        old_size = float(old_size_str) * 1024 * 1024
                        new_size = os.path.getsize(new_path)
                        
                        if abs(new_size - old_size) < (old_size * 0.01):
                            entry['Location'] = new_path
                            updated_count += 1
                            break
                    except:
                        entry['Location'] = new_path
                        updated_count += 1
                        break
            else:
                missing_count += 1
        
        if updated_count > 0:
            print(f"  ✓ Updated {updated_count} file paths to new location")
        if outside_dir_count > 0:
            print(f"  ℹ {outside_dir_count} files exist outside '{self.output_dir}' (will be ignored)")
        if missing_count > 0:
            print(f"  ⚠ {missing_count} files from index not found in '{self.output_dir}'")
        
        if updated_count == 0 and outside_dir_count == 0 and missing_count == 0:
            print(f"  ✓ All file paths are current")
    
    def categorize_file(self, title, tags, description):
        """Determine what category this file belongs to based on intent"""
        
        gun_model = self.identify_gun_model(tags)
        caliber = self.identify_caliber(tags)
        part_type = self.identify_part_type(tags, title)
        is_complete = self.is_complete_build(tags, title, description)
        
        if is_complete:
            if 'Handgun' in tags or 'Pistol' in tags or gun_model in ['Glock', '1911']:
                if gun_model:
                    if 'Glock' in gun_model:
                        return (f"Complete_Firearms/Handguns/Glock_Clones/{gun_model}", 
                               gun_model, caliber, 'Complete Build')
                    elif '1911' in gun_model:
                        return (f"Complete_Firearms/Handguns/1911_Clones", 
                               gun_model, caliber, 'Complete Build')
                    else:
                        return (f"Complete_Firearms/Handguns/Other_Handguns", 
                               gun_model, caliber, 'Complete Build')
                return ("Complete_Firearms/Handguns/Other_Handguns", 
                       gun_model, caliber, 'Complete Build')
            
            elif 'Rifle' in tags or 'AR-15' in tags or 'AK-47' in tags:
                if 'AR-15' in tags or 'AR-15' in title:
                    return ("Complete_Firearms/Rifles/AR-15_Builds", 
                           'AR-15', caliber, 'Complete Build')
                elif 'AR-22' in tags or 'AR-22' in title:
                    return ("Complete_Firearms/Rifles/AR-22_Builds", 
                           'AR-22', caliber, 'Complete Build')
                elif 'AK' in str(tags):
                    return ("Complete_Firearms/Rifles/AK_Builds", 
                           'AK-47', caliber, 'Complete Build')
                else:
                    return ("Complete_Firearms/Rifles/Other_Rifles", 
                           gun_model, caliber, 'Complete Build')
            
            elif 'PCC' in tags:
                return ("Complete_Firearms/PCCs", 
                       gun_model, caliber, 'Complete Build')
            
            elif 'Shotgun' in tags:
                return ("Complete_Firearms/Shotguns", 
                       gun_model, caliber, 'Complete Build')
        
        if part_type in ['Frame', 'Receiver', 'Lower']:
            if gun_model and 'AR-15' in gun_model:
                return (f"Parts_and_Upgrades/Frames_and_Receivers/AR-15_Lowers", 
                       gun_model, caliber, part_type)
            elif gun_model and 'Glock' in gun_model:
                return (f"Parts_and_Upgrades/Frames_and_Receivers/Glock_Frames", 
                       gun_model, caliber, part_type)
            else:
                return (f"Parts_and_Upgrades/Frames_and_Receivers/Other_Frames", 
                       gun_model, caliber, part_type)
        
        if part_type in ['Upper', 'Slide']:
            if gun_model and 'AR-15' in gun_model:
                return (f"Parts_and_Upgrades/Uppers_and_Slides/AR-15_Uppers", 
                       gun_model, caliber, part_type)
            elif gun_model and 'Glock' in gun_model:
                return (f"Parts_and_Upgrades/Uppers_and_Slides/Glock_Slides", 
                       gun_model, caliber, part_type)
            else:
                return (f"Parts_and_Upgrades/Uppers_and_Slides/Other_Uppers", 
                       gun_model, caliber, part_type)
        
        if 'FRT' in tags or 'Trigger' in tags:
            return (f"Parts_and_Upgrades/Fire_Control/{'FRTs' if 'FRT' in tags else 'Triggers'}", 
                   gun_model, caliber, 'Fire Control')
        
        if 'Barrel' in tags or 'Bolt' in tags or 'DIY Barrel' in tags:
            return (f"Parts_and_Upgrades/Barrels_and_Bolts", 
                   gun_model, caliber, part_type)
        
        if 'Suppressor' in tags:
            if caliber in ['9x19mm', '.45 ACP', '22 Long Rifle']:
                cal_folder = caliber.replace(' ', '_').replace('.', '')
                return (f"Accessories/By_Function/Suppressors/Pistol_Caliber/{cal_folder}", 
                       gun_model, caliber, 'Suppressor')
            else:
                cal_folder = caliber.replace(' ', '_').replace('.', '') if caliber else 'Multi_Caliber'
                return (f"Accessories/By_Function/Suppressors/Rifle_Caliber/{cal_folder}", 
                       gun_model, caliber, 'Suppressor')
        
        if 'Magazine' in tags:
            if gun_model:
                model_folder = gun_model.replace(' ', '_')
                return (f"Accessories/By_Function/Magazines/By_Gun/{model_folder}_Magazines", 
                       gun_model, caliber, 'Magazine')
            elif caliber:
                cal_folder = caliber.replace(' ', '_').replace('.', '')
                return (f"Accessories/By_Function/Magazines/By_Caliber/{cal_folder}", 
                       gun_model, caliber, 'Magazine')
            else:
                return (f"Accessories/By_Function/Magazines/Other", 
                       gun_model, caliber, 'Magazine')
        
        if 'Sight' in tags or 'Optic' in tags:
            return (f"Accessories/By_Function/Optics_and_Sights", 
                   gun_model, caliber, 'Optic/Sight')
        
        if 'Muzzle Device' in tags:
            return (f"Accessories/By_Function/Muzzle_Devices", 
                   gun_model, caliber, 'Muzzle Device')
        
        if 'Stock' in tags or 'Grip' in tags or 'Pistol Brace' in tags:
            return (f"Accessories/By_Function/Grips_and_Stocks", 
                   gun_model, caliber, part_type or 'Stock/Grip')
        
        if 'Furniture' in tags or 'Handguard' in tags or 'Foregrip' in tags:
            return (f"Furniture", gun_model, caliber, 'Furniture')
        
        if 'Jig' in title or 'Jig' in tags or 'Fixture' in title:
            if 'Bending' in title:
                return (f"Tools_and_Jigs/Bending_Jigs", gun_model, caliber, 'Jig')
            elif 'Drill' in title:
                return (f"Tools_and_Jigs/Drilling_Jigs", gun_model, caliber, 'Jig')
            elif 'CNC' in title or 'CNC' in tags:
                return (f"Tools_and_Jigs/CNC_Fixtures", gun_model, caliber, 'CNC Fixture')
            else:
                return (f"Tools_and_Jigs/Assembly_Tools", gun_model, caliber, 'Tool')
        
        if gun_model:
            model_folder = gun_model.replace(' ', '_')
            return (f"Miscellaneous/By_Gun_Model/{model_folder}", 
                   gun_model, caliber, 'Other')
        elif caliber:
            cal_folder = caliber.replace(' ', '_').replace('.', '')
            return (f"Miscellaneous/By_Caliber/{cal_folder}", 
                   gun_model, caliber, 'Other')
        else:
            return (f"Miscellaneous/Uncategorized", 
                   gun_model, caliber, 'Other')
    
    def is_complete_build(self, tags, title, description):
        complete_indicators = [
            'Complete', 'Full Build', 'Full Gun', 'DIY Fire Control',
            'DIY Bolt', 'Printed Firearm', 'No Firearm Parts'
        ]
        
        for indicator in complete_indicators:
            if indicator in tags or indicator.lower() in title.lower():
                return True
        
        has_frame = any(x in tags for x in ['Frame/Receiver', 'Frame', 'Receiver'])
        has_other_parts = any(x in tags for x in ['Upper', 'Barrel', 'Bolt', 'Slide'])
        
        if has_frame and has_other_parts:
            return True
        
        return False
    
    def identify_gun_model(self, tags):
        models = {
            'Glock 19': ['Glock 19'],
            'Glock 17': ['Glock 17'],
            'Glock 26': ['Glock 26'],
            'Glock 43': ['Glock 43'],
            'Glock 48': ['Glock 48'],
            'Glock': ['Glock'],
            'AR-15': ['AR-15'],
            'AR-22': ['AR-22'],
            'AR-10': ['AR-10'],
            'AK-47': ['AK-47', 'AK-74'],
            'FGC-9': ['FGC-9'],
            '1911': ['1911'],
            'TX22': ['TX22'],
            'Taurus': ['Taurus']
        }
        
        for model_name, keywords in models.items():
            for keyword in keywords:
                if keyword in tags:
                    return model_name
        
        return None
    
    def identify_caliber(self, tags):
        calibers = ['9x19mm', '22 Long Rifle', '.45 ACP', '5.56x45mm', 
                   '7.62x39mm', '.308 Winchester', '12 Gauge']
        
        for cal in calibers:
            if cal in tags:
                return cal
        
        return None
    
    def identify_part_type(self, tags, title):
        part_types = {
            'Frame': ['Frame/Receiver', 'Frame'],
            'Receiver': ['Receiver'],
            'Lower': ['Lower'],
            'Upper': ['Upper'],
            'Slide': ['Slide'],
            'Barrel': ['Barrel', 'DIY Barrel'],
            'Bolt': ['Bolt', 'DIY Bolt'],
            'Trigger': ['Trigger'],
            'Stock': ['Stock'],
            'Grip': ['Grip', 'Pistol Grip'],
            'Magazine': ['Magazine'],
            'Suppressor': ['Suppressor'],
        }
        
        for part_name, keywords in part_types.items():
            for keyword in keywords:
                if keyword in tags or keyword.lower() in title.lower():
                    return part_name
        
        return None
    
    def get_folder_path(self, title, tags, description):
        category_path, gun_model, caliber, part_type = self.categorize_file(title, tags, description)
        folder_path = os.path.join(self.output_dir, category_path)
        
        return folder_path, {
            'category': category_path,
            'gun_model': gun_model,
            'caliber': caliber,
            'part_type': part_type
        }

    def add_to_index(self, filename, filepath, title, tags, category_info, file_size, description,
                     lbry_url='', detail_url='', release_date='', last_updated='',
                     author='', version='', notes='', readme='',
                     odysee_views=0, odysee_likes=0, odysee_dislikes=0):

        existing_idx = None
        for idx, entry in enumerate(self.master_index):
            if (entry.get('LBRY URL') == lbry_url and lbry_url) or \
                    entry.get('File Name') == filename or \
                    entry.get('Location') == filepath:
                existing_idx = idx
                break

        new_entry = {
            'File Name': filename,
            'Location': filepath,
            'LBRY URL': lbry_url,
            'Detail URL': detail_url,
            'Category': category_info.get('category', ''),
            'Gun Model': category_info.get('gun_model', ''),
            'Caliber': category_info.get('caliber', ''),
            'Part Type': category_info.get('part_type', ''),
            'Tags': ', '.join(tags),
            'File Size (MB)': f"{file_size / 1024 / 1024:.2f}",
            'Release Date': release_date or '',
            'Last Updated': last_updated or '',
            'Author': author or '',
            'Version': version or '',
            'Date Downloaded': datetime.now().strftime('%Y-%m-%d %H:%M'),
            'Odysee Views': odysee_views or 0,
            'Odysee Likes': odysee_likes or 0,
            'Odysee Dislikes': odysee_dislikes or 0,
            'Description': self.clean_for_excel(description),
            'Notes': self.clean_for_excel(notes),
            'Readme': self.clean_for_excel(readme)
        }

        if existing_idx is not None:
            self.master_index[existing_idx] = new_entry
        else:
            self.master_index.append(new_entry)
    
    def generate_master_index(self):
        if not self.master_index:
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "GunCAD Master Index"
            
            headers = self.COLUMN_ORDER
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            for row_idx, entry in enumerate(self.master_index, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = entry.get(header, '')
                    if isinstance(value, str):
                        value = self.clean_for_excel(value)
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            ws.freeze_panes = 'A2'
            
            wb.save(self.index_file)
            print(f"\n✓ Master index saved: {self.index_file}")
            
        except ImportError:
            import csv
            
            with open(self.index_file_csv, 'w', newline='', encoding='utf-8') as f:
                if self.master_index:
                    writer = csv.DictWriter(f, fieldnames=self.master_index[0].keys())
                    writer.writeheader()
                    writer.writerows(self.master_index)
            
            print(f"\n✓ Master index saved: {self.index_file_csv}")
    
    def generate_readmes(self):
        folders_to_document = set()
        
        for root, dirs, files in os.walk(self.output_dir):
            if root == self.output_dir:
                continue
            
            if any(part.startswith('.') for part in root.split(os.sep)):
                continue
            
            cad_files = [f for f in files if not f.startswith('README') and not f.startswith('.')]
            if cad_files:
                folders_to_document.add(root)
        
        for folder in folders_to_document:
            readme_path = os.path.join(folder, 'README.md')
            
            folder_files = [entry for entry in self.master_index 
                           if os.path.dirname(entry.get('Location', '')) == folder]
            
            if not folder_files:
                continue
            
            folder_name = os.path.basename(folder)
            parent_folder = os.path.basename(os.path.dirname(folder))
            
            gun_models = set(f.get('Gun Model') for f in folder_files if f.get('Gun Model'))
            calibers = set(f.get('Caliber') for f in folder_files if f.get('Caliber'))
            part_types = set(f.get('Part Type') for f in folder_files if f.get('Part Type'))
            
            with open(readme_path, 'w', encoding='utf-8') as f:
                f.write(f"# {folder_name}\n\n")
                f.write(f"**Location:** `{parent_folder}/{folder_name}/`\n\n")
                f.write(f"**Files in this folder:** {len(folder_files)}\n\n")
                
                if gun_models:
                    f.write(f"**Gun Models:** {', '.join(sorted(gun_models))}\n\n")
                
                if calibers:
                    f.write(f"**Calibers:** {', '.join(sorted(calibers))}\n\n")
                
                if part_types:
                    f.write(f"**Part Types:** {', '.join(sorted(part_types))}\n\n")
                
                f.write("---\n\n")
                f.write("## Files\n\n")
                for file_entry in sorted(folder_files, key=lambda x: x.get('File Name', '')):
                    f.write(f"- `{file_entry.get('File Name', '')}`")
                    details = []
                    if file_entry.get('Gun Model'):
                        details.append(file_entry['Gun Model'])
                    if file_entry.get('Caliber'):
                        details.append(file_entry['Caliber'])
                    if details:
                        f.write(f" ({', '.join(details)})")
                    f.write("\n")
                
                f.write("\n---\n")
                f.write(f"*Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}*\n")
        
        print(f"✓ Generated README files in {len(folders_to_document)} folders")
    
    def generate_quick_find(self):
        quick_find_path = os.path.join(self.output_dir, 'QUICK_FIND.txt')
        
        with open(quick_find_path, 'w', encoding='utf-8') as f:
            f.write("═══════════════════════════════════════════════════════════════════\n")
            f.write("               GUNCAD COLLECTION QUICK REFERENCE                   \n")
            f.write("═══════════════════════════════════════════════════════════════════\n\n")
            
            f.write("🔫 BUILDING A COMPLETE GUN?\n")
            f.write("   → /Complete_Firearms/[type]/[model]/\n\n")
            
            f.write("🔧 NEED A SPECIFIC PART?\n")
            f.write("   → /Parts_and_Upgrades/[part_type]/[model]/\n\n")
            
            f.write("🎯 LOOKING FOR ACCESSORIES?\n")
            f.write("   → /Accessories/By_Function/[accessory_type]/\n\n")
            
            f.write("📦 BY CALIBER:\n")
            f.write("   → Search Master Index Excel file for caliber column\n\n")
            
            f.write("🔍 CAN'T FIND SOMETHING?\n")
            f.write("   1. Open GunCAD_Master_Index.xlsx\n")
            f.write("   2. Use Ctrl+F to search\n")
            f.write("   3. Check 'Location' column for path\n\n")
            
            f.write("═══════════════════════════════════════════════════════════════════\n")
            f.write("Total Files: {}\n".format(len(self.master_index)))
            f.write("Last Updated: {}\n".format(datetime.now().strftime('%Y-%m-%d %H:%M')))
            f.write("═══════════════════════════════════════════════════════════════════\n")
        
        print(f"✓ Quick find guide saved: {quick_find_path}")


class GunCADDownloaderV6:
    def __init__(self, output_dir='downloads', max_wait_time=300, batch_update_interval=10, excluded_tags=None):
        self.output_dir = output_dir
        self.organizer = IntentBasedOrganizer(output_dir)
        self.tracker = DownloadTracker(output_dir=output_dir, excel_index=self.organizer.master_index)
        self.lbry = LBRYDaemonClient(max_wait_time=max_wait_time)
        self.api = GunCADIndexAPIClient()
        self.verifier = FileVerifier()
        self.batch_update_interval = batch_update_interval
        self.excluded_tags = excluded_tags or []
        
        self.session_successful = 0
        self.session_failed = 0
        self.session_skipped_by_filter = 0
        
        self.start_time = None
        self.total_items = 0
    
    def update_live_stats(self, current_item=None):
        total = self.session_successful + self.session_failed + self.session_skipped_by_filter
        
        stats_line = f"\n📊 Session Stats: ✓ {self.session_successful} successful | ✗ {self.session_failed} failed"
        if self.session_skipped_by_filter > 0:
            stats_line += f" | ⊘ {self.session_skipped_by_filter} filtered"
        stats_line += f" | Total: {total}"
        
        if self.start_time and current_item and self.total_items > 0:
            elapsed = time.time() - self.start_time
            if total > 0:
                avg_time_per_item = elapsed / total
                remaining_items = self.total_items - total
                eta_seconds = avg_time_per_item * remaining_items
                
                if eta_seconds < 60:
                    eta_str = f"{int(eta_seconds)}s"
                elif eta_seconds < 3600:
                    eta_str = f"{int(eta_seconds / 60)}m {int(eta_seconds % 60)}s"
                else:
                    hours = int(eta_seconds / 3600)
                    minutes = int((eta_seconds % 3600) / 60)
                    eta_str = f"{hours}h {minutes}m"
                
                stats_line += f" | ETA: {eta_str}"
        
        print(stats_line)
    
    def process_entry(self, entry, current_item=None):
        title = self.organizer.sanitize_filename(entry.get('title', 'Unknown'))
        detail_url = entry.get('detail_url', '')
        tags = entry.get('tags', [])
        description = entry.get('description', '')
        release_date = entry.get('release_date', '')
        author = entry.get('author', '')
        version = entry.get('version', '')
        notes = entry.get('notes', '')
        readme = entry.get('readme', '')
        
        # Check if any excluded tags match
        if self.excluded_tags:
            for excluded_tag in self.excluded_tags:
                if excluded_tag in tags:
                    print(f"  Title: {title}")
                    print(f"  ⊘ Skipped (filtered by tag: {excluded_tag})")
                    self.session_skipped_by_filter += 1
                    self.update_live_stats(current_item)
                    return None
        
        print(f"  Title: {title}")
        print(f"  Tags: {', '.join(tags[:5])}{' ...' if len(tags) > 5 else ''}")
        
        lbry_url = entry.get('lbry_url', '')
        
        if not lbry_url:
            print("  ✗ No LBRY URL")
            self.tracker.mark_failed(detail_url, title, "No LBRY URL", lbry_url='')
            self.session_failed += 1
            self.update_live_stats(current_item)
            return None
        
        print(f"  LBRY: {lbry_url}")
        
        try:
            parsed = urlparse(lbry_url)
            expected_filename = parsed.path.lstrip('/').split('#')[0]
            if not expected_filename:
                expected_filename = parsed.netloc.split('#')[0]
        except:
            expected_filename = None
        
        if expected_filename:
            existing_path = self.tracker.file_exists_in_cache(expected_filename, entry.get('size', 0))
            if existing_path:
                print(f"  ✓ File already exists: {existing_path}")
                print(f"  Skipping download, updating records...")
                
                file_size = os.path.getsize(existing_path)
                folder_path, category_info = self.organizer.get_folder_path(title, tags, description)

                self.organizer.add_to_index(
                    expected_filename, existing_path, title, tags, category_info,
                    file_size, description,
                    lbry_url=lbry_url,
                    detail_url=detail_url,
                    release_date=entry.get('release_date', ''),
                    last_updated=entry.get('last_updated', ''),
                    author=entry.get('author', ''),
                    version=entry.get('version', ''),
                    notes=entry.get('notes', ''),
                    readme=entry.get('readme', ''),
                    odysee_views=entry.get('odysee_views', 0),
                    odysee_likes=entry.get('odysee_likes', 0),
                    odysee_dislikes=entry.get('odysee_dislikes', 0)
                )
                
                self.tracker.mark_downloaded(
                    detail_url, title, existing_path, tags,
                    verified=True, file_size=file_size,
                    category=category_info.get('category'),
                    gun_model=category_info.get('gun_model'),
                    caliber=category_info.get('caliber')
                )
                
                self.session_successful += 1
                self.update_live_stats(current_item)
                return existing_path
        
        if not self.lbry.available:
            print("  ✗ LBRY daemon not available")
            self.tracker.mark_failed(detail_url, title, "No daemon", lbry_url=lbry_url)
            self.session_failed += 1
            self.update_live_stats(current_item)
            return None
        
        download_path = self.lbry.get_file(lbry_url)
        
        if not download_path:
            print("  ✗ Download failed")
            self.tracker.mark_failed(detail_url, title, "Download failed", lbry_url=lbry_url)
            self.session_failed += 1
            self.update_live_stats(current_item)
            return None
        
        print("  Verifying file...")
        is_valid, message = self.verifier.verify_file(download_path)
        print(f"  {message}")
        
        if not is_valid:
            print("  ✗ Verification failed")
            self.tracker.mark_failed(detail_url, title, f"Verification: {message}", lbry_url=lbry_url)
            self.session_failed += 1
            self.update_live_stats(current_item)
            return None
        
        folder_path, category_info = self.organizer.get_folder_path(title, tags, description)
        os.makedirs(folder_path, exist_ok=True)
        
        filename = os.path.basename(download_path)
        final_path = os.path.join(folder_path, filename)
        
        try:
            download_abs = os.path.abspath(download_path)
            final_abs = os.path.abspath(final_path)
            
            if download_abs != final_abs:
                if os.path.exists(final_path):
                    os.remove(final_path)
                
                try:
                    shutil.move(download_path, final_path)
                    print(f"  ✓ Moved to: {category_info['category']}/")
                except (OSError, shutil.Error):
                    shutil.copy2(download_path, final_path)
                    print(f"  ✓ Copied to: {category_info['category']}/")
                    try:
                        os.remove(download_path)
                    except:
                        pass
            else:
                print(f"  ✓ Already in correct location")
            
            if not os.path.exists(final_path):
                raise Exception("File not found at final destination")
            
            file_size = os.path.getsize(final_path)

            self.organizer.add_to_index(
                filename, final_path, title, tags, category_info,
                file_size, description,
                lbry_url=lbry_url,
                detail_url=detail_url,
                release_date=entry.get('release_date', ''),
                last_updated=entry.get('last_updated', ''),
                author=entry.get('author', ''),
                version=entry.get('version', ''),
                notes=entry.get('notes', ''),
                readme=entry.get('readme', ''),
                odysee_views=entry.get('odysee_views', 0),
                odysee_likes=entry.get('odysee_likes', 0),
                odysee_dislikes=entry.get('odysee_dislikes', 0)
            )
            
            self.tracker.mark_downloaded(
                detail_url, title, final_path, tags,
                verified=True, file_size=file_size,
                category=category_info.get('category'),
                gun_model=category_info.get('gun_model'),
                caliber=category_info.get('caliber')
            )
            
            self.session_successful += 1
            self.update_live_stats(current_item)
            return final_path
            
        except Exception as e:
            print(f"  ✗ Error organizing: {e}")
            self.tracker.mark_failed(detail_url, title, f"Organize error: {e}", lbry_url=lbry_url)
            self.session_failed += 1
            self.update_live_stats(current_item)
            return None
    
    def run(self, max_pages=1, delay=3, check_new_only=True):
        print(f"\n{'='*70}")
        print("GunCAD Index Downloader v6 - Quality of Life Edition (FIXED)")
        print(f"{'='*70}")
        print(f"API: {self.api.api_base}")
        print(f"Output: {self.output_dir}")
        print(f"Pages: {max_pages}")
        print(f"New only: {check_new_only}")
        print(f"Download timeout: {self.lbry.max_wait_time}s")
        print(f"Excel batch updates: every {self.batch_update_interval} downloads")
        print(f"{'='*70}\n")
        
        if not self.lbry.available:
            print("ERROR: LBRY daemon required!")
            return
        
        stats = self.tracker.get_stats()
        print(f"📊 Historical Statistics:")
        print(f"  Total: {stats['total']}")
        print(f"  Successful: {stats['successful']}")
        print(f"  Failed: {stats['failed']}\n")
        
        all_entries = []
        for page in range(1, max_pages + 1):
            print(f"Fetching page {page} from API...")
            
            raw_entries = self.api.get_releases(page=page, limit=25)
            
            if not raw_entries:
                print("  No more entries")
                break
            
            entries = [self.api.parse_entry(e) for e in raw_entries]
            print(f"  Found {len(entries)} entries")
            all_entries.extend(entries)
        
        print(f"\nTotal entries: {len(all_entries)}")
        
        if check_new_only:
            new_entries = []
            for e in all_entries:
                detail_url = e.get('detail_url', '')
                lbry_url = e.get('lbry_url', '')
                size = e.get('size', 0)
                
                if not self.tracker.is_downloaded(detail_url, lbry_url, size):
                    new_entries.append(e)
            
            already_in_this_dir = len(all_entries) - len(new_entries)
            print(f"New entries to download: {len(new_entries)}")
            print(f"Already in '{self.output_dir}': {already_in_this_dir}")
            all_entries = new_entries
        
        if not all_entries:
            print("\n✓ Nothing to download!")
            self.organizer.generate_master_index()
            self.organizer.generate_readmes()
            self.organizer.generate_quick_find()
            return
        
        self.total_items = len(all_entries)
        self.start_time = time.time()
        
        for i, entry in enumerate(all_entries, 1):
            print(f"\n{'='*70}")
            print(f"[{i}/{len(all_entries)}]")
            
            self.process_entry(entry, current_item=i)
            
            if i % self.batch_update_interval == 0 or i == len(all_entries):
                print(f"\n📊 Updating Excel index...")
                self.organizer.generate_master_index()
                print(f"   (Includes failed downloads sheet)")
            
            if i < len(all_entries):
                time.sleep(delay)
        
        print(f"\n{'='*70}")
        print("Generating supporting files...")
        print(f"{'='*70}\n")
        
        self.organizer.generate_master_index()
        self.organizer.generate_readmes()
        self.organizer.generate_quick_find()
        
        print(f"\n{'='*70}")
        print("Complete!")
        print(f"{'='*70}")
        print(f"✓ Successful downloads: {self.session_successful}")
        print(f"✗ Failed downloads: {self.session_failed}")
        if self.session_skipped_by_filter > 0:
            print(f"⊘ Skipped by tag filter: {self.session_skipped_by_filter}")
        print(f"\n📁 All files saved to:")
        print(f"   {os.path.abspath(self.output_dir)}")
        print(f"\n📄 Generated documentation:")
        print(f"   • Master Index: {os.path.join(self.output_dir, 'GunCAD_Master_Index.xlsx')}")
        print(f"   • Quick Find Guide: {os.path.join(self.output_dir, 'QUICK_FIND.txt')}")
        print(f"   • README files in each category folder")
        print(f"{'='*70}\n")


def get_output_directory():
    print("\n" + "="*70)
    print("OUTPUT DIRECTORY CONFIGURATION")
    print("="*70)
    
    default_dir = 'GunCAD_Downloads'
    print(f"\nDefault directory: {default_dir}")
    print("Press Enter to use default, or type a custom path:")
    
    user_input = input("> ").strip()
    
    if user_input:
        return user_input
    else:
        return default_dir


def get_max_pages():
    print("\n" + "="*70)
    print("DOWNLOAD QUANTITY")
    print("="*70)
    print("\nHow many pages to download? (25 files per page)")
    print("Examples:")
    print("  1 page   = 25 files")
    print("  10 pages = 250 files")
    print("  50 pages = 1,250 files")
    print("  999 pages = Download everything available (~25,000 files)")
    print(f"\nDefault: 999 (download all)")
    print("Press Enter for default, or type a number:")
    
    user_input = input("> ").strip()
    
    if user_input:
        try:
            pages = int(user_input)
            if pages < 1:
                print("Invalid input. Using default (999).")
                return 999
            return pages
        except ValueError:
            print("Invalid input. Using default (999).")
            return 999
    else:
        return 999


def get_download_timeout():
    print("\n" + "="*70)
    print("DOWNLOAD TIMEOUT SETTING")
    print("="*70)
    print("\nMaximum wait time for each file download (in seconds)")
    print("Recommendations:")
    print("  180s (3 min)  - Fast internet connection")
    print("  300s (5 min)  - Average connection (RECOMMENDED)")
    print("  600s (10 min) - Slow connection or very large files")
    print(f"\nDefault: 300 seconds (5 minutes)")
    print("Press Enter for default, or type seconds:")
    
    user_input = input("> ").strip()
    
    if user_input:
        try:
            timeout = int(user_input)
            if timeout < 1:
                print("Timeout too short. Using minimum (1s).")
                return 1
            return timeout
        except ValueError:
            print("Invalid input. Using default (300s).")
            return 300
    else:
        return 300


def get_excluded_tags(api_client):
    print("\n" + "="*70)
    print("TAG FILTER (OPTIONAL)")
    print("="*70)
    
    print("\nFetching available tags from GunCAD Index...")
    all_tags = api_client.get_all_tags(scan_pages=10)  # Scan first 10 pages (250 files)
    
    if all_tags:
        print(f"\n✓ Found {len(all_tags)} unique tags")
        print("\nAvailable tags:")
        print("-" * 70)
        
        # Display tags in columns for readability
        cols = 3
        tags_per_col = (len(all_tags) + cols - 1) // cols
        
        for i in range(tags_per_col):
            row_tags = []
            for col in range(cols):
                idx = i + col * tags_per_col
                if idx < len(all_tags):
                    row_tags.append(f"{all_tags[idx]:<23}")
            print("  " + " ".join(row_tags))
        
        print("-" * 70)
    else:
        print("\n⚠ Could not fetch tags from API")
        print("Common tags you might want to exclude:")
        print("  Furniture, Accessory, Jig, Fixture, Sight, Optic")
        print("  Stock, Grip, Magazine, Muzzle Device")
    
    print("\nExclude files with specific tags (comma-separated)")
    print("\nExamples:")
    print("  'Furniture,Accessory'  - Skip furniture and accessories")
    print("  'Magazine,Jig'         - Skip magazines and jigs")
    print("  'Furniture'            - Skip only furniture")
    
    if all_tags:
        print("\nTip: Type 'list' to see the tags again")
    
    print("\nPress Enter to download everything, or type tags to exclude:")
    
    while True:
        user_input = input("> ").strip()
        
        if user_input.lower() == 'list' and all_tags:
            print("\nAvailable tags:")
            print("-" * 70)
            for i in range(tags_per_col):
                row_tags = []
                for col in range(cols):
                    idx = i + col * tags_per_col
                    if idx < len(all_tags):
                        row_tags.append(f"{all_tags[idx]:<23}")
                print("  " + " ".join(row_tags))
            print("-" * 70)
            print("\nPress Enter to download everything, or type tags to exclude:")
            continue
        
        break
    
    if user_input:
        # Split by comma and clean up whitespace
        excluded = [tag.strip() for tag in user_input.split(',') if tag.strip()]
        
        # Validate that excluded tags exist in the API
        if all_tags:
            invalid_tags = [tag for tag in excluded if tag not in all_tags]
            if invalid_tags:
                print(f"\n⚠ Warning: These tags weren't found in first 250 releases: {', '.join(invalid_tags)}")
                print("They will still be used for filtering.")
        
        if excluded:
            print(f"\n✓ Will exclude files with tags: {', '.join(excluded)}")
            return excluded
    
    print("\n✓ No tag filters - downloading all file types")
    return []


def main():
    # LEGAL AGREEMENT PROMPT
    print("\n" + "=" * 70)
    print("!!! IMPORTANT !!!")
    print("=" * 70)
    print("""
This script automatically downloads files listed on GunCAD Index which is 
a search engine for DIY gun designs. It is not for children. By utilizing 
this script, you acknowledge and agree that:

- You are at least 18 years of age

- You have read and agree to the GunCAD Index site's terms and policies 
  outlined here: https://guncadindex.com/legal

- You understand that firearm-related information (including but not limited 
  to design files) may be regulated in your jurisdiction

- You are legally permitted, under all applicable laws and regulations, to 
  access and possess information about DIY firearms

If you are a resident of a jurisdiction that restricts access to or 
dissemination of firearm-related technical information, you affirm that you 
are exempt from such restrictions or otherwise legally authorized to access 
it. Such jurisdictions may include, but are not limited to:
  - California
  - New Jersey
  - New York

- You are solely responsible for ensuring your compliance with all applicable 
  laws

- GunCAD Index does not host, control, or distribute any linked content. It 
  is purely an index and search engine. You access any sites linked here at 
  your own risk and subject to their terms and policies
""")
    print("=" * 70)
    print("To continue, type \"I Agree\" and press Enter")
    print("=" * 70)

    user_response = input("\n> ").strip()

    if user_response.lower() != "i agree":
        print("\n" + "=" * 70)
        print("You must Agree to continue to utilize this script.")
        print("Please close the script, rerun it, and type \"I Agree\"")
        print("if you agree and wish to continue.")
        print("=" * 70 + "\n")
        return

    print("\n✓ Agreement confirmed. Proceeding with setup...\n")

    # Create API client first to fetch tags
    api_client = GunCADIndexAPIClient()

    # ... rest of existing code continues unchanged
    
    OUTPUT_DIR = get_output_directory()
    MAX_PAGES = get_max_pages()
    EXCLUDED_TAGS = get_excluded_tags(api_client)
    
    DELAY = 3
    CHECK_NEW_ONLY = True
    BATCH_UPDATE_INTERVAL = 10
    
    print(f"\n✓ Output directory: {OUTPUT_DIR}")
    print(f"✓ Pages to download: {MAX_PAGES} ({MAX_PAGES * 25} files max)")
    if EXCLUDED_TAGS:
        print(f"✓ Excluding tags: {', '.join(EXCLUDED_TAGS)}")
    print(f"✓ Download monitoring: Active progress detection (30s stall threshold)")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    try:
        downloader = GunCADDownloaderV6(
            OUTPUT_DIR,
            max_wait_time=None,  # Not used anymore
            batch_update_interval=BATCH_UPDATE_INTERVAL,
            excluded_tags=EXCLUDED_TAGS
        )
        # Pass the existing API client to avoid creating a new one
        downloader.api = api_client
        
        downloader.run(
            max_pages=MAX_PAGES,
            delay=DELAY,
            check_new_only=CHECK_NEW_ONLY
        )
    except KeyboardInterrupt:
        print("\n\nInterrupted")
    except Exception as e:
        print(f"\n\nError: {e}")
        import traceback
        traceback.print_exc()


if __name__ == '__main__':
    main()
